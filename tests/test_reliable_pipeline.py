from pathlib import Path

from openpyxl import load_workbook

from extraction.address import extract_postcode
from extraction.assets import classify_candidates, normalize_url
from extraction.models import AssetCandidate, AssetType, Property
from extraction.pipeline import _address_retry_candidates
from extraction.schema import COLUMNS, normalize_record
from extraction.validation import validate_property
from spreadsheet import LINE_HEIGHT, QA_COLUMNS, _estimate_wrapped_lines, write_xlsx


def test_postcode_extraction_and_normalisation():
    assert extract_postcode("18 Bevis Marks, London ec3a7jb") == "EC3A 7JB"
    assert extract_postcode("No postcode supplied") == ""


def test_address_retry_candidates_are_ordered_and_unique():
    assert _address_retry_candidates("Princes House, 38 Jermyn Street, SW1Y", None) == [
        "Princes House, 38 Jermyn Street",
        "38 Jermyn Street, SW1Y",
        "38 Jermyn Street",
    ]


def test_asset_classification_and_tracking_parameter_deduplication():
    candidates = classify_candidates(
        [
            AssetCandidate("https://EXAMPLE.com/file.pdf?utm_source=email", "html", anchor_text="View brochure"),
            AssetCandidate("https://example.com/file.pdf", "html", anchor_text="Duplicate brochure"),
            AssetCandidate("https://example.com/third-floor-plan.png", "html", alt_text="Floor plan"),
            AssetCandidate("https://example.com/reception.jpg", "html", alt_text="Reception"),
            AssetCandidate("https://example.com/company-logo.png", "html", alt_text="Company logo"),
            AssetCandidate("https://example.com/open-tracking.gif", "html", alt_text="tracking pixel"),
        ]
    )
    assert [item.classification for item in candidates] == [
        AssetType.BROCHURE,
        AssetType.FLOORPLAN,
        AssetType.PROPERTY_IMAGE,
        AssetType.LOGO,
        AssetType.DECORATIVE,
    ]
    assert normalize_url("javascript:alert(1)") == ""


def test_validation_generates_issues_without_destroying_values():
    record = normalize_record(
        {
            "Building": "Example House",
            "Brochure PDF": "https://example.com/photo.jpg",
            "Floor Plan": "https://example.com/shared.jpg",
            "High Res Images": "https://example.com/shared.jpg",
        }
    )
    prop = validate_property(Property.from_record(record, "input.eml", "Example", "rule:test"))
    assert prop.values["Brochure PDF"] == "https://example.com/photo.jpg"
    assert prop.review_required
    assert {issue.field for issue in prop.issues} >= {"Brochure PDF", "High Res Images"}


def test_missing_optional_data_is_not_an_error():
    record = normalize_record({"Building": "18 Bevis Marks, London EC3A 7JB"})
    prop = validate_property(Property.from_record(record, "input.pdf", "Example", "rule:test"))
    assert not any(issue.field in {"Brochure PDF", "Floor Plan", "High Res Images"} for issue in prop.issues)


def test_spreadsheet_named_mapping_hyperlinks_and_qa_sheet(tmp_path: Path):
    record = normalize_record({"Building": "Example House", "Brochure PDF": "https://example.com/brochure.pdf"})
    prop = validate_property(Property.from_record(record, "source.eml", "Example", "rule:test"))
    path = tmp_path / "result.xlsx"
    write_xlsx(path, [prop.to_record()], include_qa_sheet=True)
    workbook = load_workbook(path)
    listings = workbook["Listings"]
    headers = [cell.value for cell in listings[1]]
    assert headers == COLUMNS
    brochure_cell = listings.cell(2, COLUMNS.index("Brochure PDF") + 1)
    assert brochure_cell.hyperlink.target == "https://example.com/brochure.pdf"
    assert [cell.value for cell in workbook["QA Review"][1]] == QA_COLUMNS


def test_spreadsheet_wraps_all_columns_and_grows_row_height(tmp_path: Path):
    long_building = "A very long building name that should wrap within the capped column width instead of spilling sideways into neighbouring cells"
    long_issue = (
        "This validation issue message is intentionally long enough that the QA Review "
        "sheet must estimate a multi-line row height once wrap_text is applied."
    )
    record = normalize_record(
        {
            "Building": long_building,
            "Special Features": "Short feature",
            "Brochure PDF": "https://example.com/brochure.pdf",
            "_validation_issues": [
                {
                    "field": "Building",
                    "message": long_issue,
                    "severity": "warning",
                    "value": long_building,
                    "action": "Confirm the building name against the source document.",
                }
            ],
        }
    )
    path = tmp_path / "wrap.xlsx"
    write_xlsx(path, [record], include_qa_sheet=True)
    workbook = load_workbook(path)
    listings = workbook["Listings"]

    for col_idx in range(1, len(COLUMNS) + 1):
        cell = listings.cell(2, col_idx)
        assert cell.alignment.wrap_text is True
        assert cell.alignment.vertical == "center"
        assert listings.column_dimensions[cell.column_letter].width <= 45

    brochure_cell = listings.cell(2, COLUMNS.index("Brochure PDF") + 1)
    assert brochure_cell.value == "Here"
    assert brochure_cell.hyperlink.target == "https://example.com/brochure.pdf"

    building_letter = listings.cell(2, COLUMNS.index("Building") + 1).column_letter
    building_width = listings.column_dimensions[building_letter].width or 10
    expected_lines = _estimate_wrapped_lines(long_building, building_width)
    assert expected_lines > 1
    assert listings.row_dimensions[2].height == expected_lines * LINE_HEIGHT

    qa = workbook["QA Review"]
    issue_cell = qa.cell(2, QA_COLUMNS.index("Issue") + 1)
    assert issue_cell.alignment.wrap_text is True
    issue_letter = issue_cell.column_letter
    issue_width = qa.column_dimensions[issue_letter].width or 10
    qa_lines = _estimate_wrapped_lines(long_issue, issue_width)
    assert qa_lines > 1
    assert qa.row_dimensions[2].height >= qa_lines * LINE_HEIGHT
    assert qa.row_dimensions[2].height > LINE_HEIGHT


def test_source_reference_survives_model_validation_and_spreadsheet_export(tmp_path: Path):
    source_url = "https://files.example.test/batches/42/source.eml"
    record = normalize_record(
        {
            "Building": "Example House",
            "Brochure PDF": "https://assets.example.test/brochure.pdf",
            "Floor Plan": "https://assets.example.test/floorplan.png",
            "High Res Images": "https://assets.example.test/reception.jpg",
        }
    )
    prop = validate_property(
        Property.from_record(record, "source.eml", "Example", "rule:test", source_url, True)
    )
    exported = prop.to_record()
    assert prop.source_file_name == "source.eml"
    assert prop.source_file_url == source_url
    assert "Link to File" not in COLUMNS
    assert "Link to File" not in exported
    assert exported["_source_file_url"] == source_url
    assert exported["_source_file_name"] == "source.eml"
    assert exported["_source_file_url"] not in {
        exported["Brochure PDF"], exported["Floor Plan"], exported["High Res Images"]
    }

    path = tmp_path / "source-link.xlsx"
    write_xlsx(path, [exported])
    headers = [cell.value for cell in load_workbook(path)["Listings"][1]]
    assert "Link to File" not in headers
    assert headers == COLUMNS


def test_missing_source_url_only_warns_when_one_was_expected():
    record = normalize_record({"Building": "Example House"})
    local = validate_property(Property.from_record(record, "local.pdf", "Example", "rule:test"))
    expected = validate_property(
        Property.from_record(record, "hosted.pdf", "Example", "rule:test", source_url_expected=True)
    )
    assert not any(issue.field == "Source file" for issue in local.issues)
    assert any(issue.field == "Source file" for issue in expected.issues)


def test_brochure_source_pdf_seeds_blank_brochure_pdf(tmp_path: Path):
    import app as app_module

    source_url = "https://files.example.test/batches/1/Clerkenwell%20Brochure.pdf"
    brochure_pdf = tmp_path / "2nd Floor - 2-7 Clerkenwell Green Brochure.pdf"
    brochure_pdf.write_bytes(b"%PDF-1.4")

    # Single-listing brochure via LLM: blank Brochure PDF gets the hosted source.
    records = [{"Building": "2-7 Clerkenwell Green", "Brochure PDF": ""}]
    app_module._seed_brochure_from_source_pdf(
        records, brochure_pdf, source_url, "llm", brochure_pdf.name
    )
    assert records[0]["Brochure PDF"] == source_url

    # Existing Brochure PDF must not be overwritten.
    records = [{"Building": "2-7 Clerkenwell Green", "Brochure PDF": "https://cdn.example/keep.pdf"}]
    app_module._seed_brochure_from_source_pdf(
        records, brochure_pdf, source_url, "llm", brochure_pdf.name
    )
    assert records[0]["Brochure PDF"] == "https://cdn.example/keep.pdf"

    # BC Current Availability is the brochure PDF for every listing row.
    schedule = tmp_path / "BC Current Availability.pdf"
    schedule.write_bytes(b"%PDF-1.4")
    records = [
        {"Building": "10-12 Alie Street", "Brochure PDF": ""},
        {"Building": "Other Building", "Brochure PDF": ""},
    ]
    app_module._seed_brochure_from_source_pdf(
        records, schedule, source_url, "rule:BC", schedule.name
    )
    assert all(r["Brochure PDF"] == source_url for r in records)

    # Email/xlsx-shaped paths never seed (suffix check).
    eml = tmp_path / "note.eml"
    eml.write_text("From: x")
    records = [{"Building": "Example", "Brochure PDF": ""}]
    app_module._seed_brochure_from_source_pdf(records, eml, source_url, "llm", eml.name)
    assert records[0]["Brochure PDF"] == ""

    # Breezblok single-property PDF seeds from source.
    breez = tmp_path / "John Stow House.pdf"
    breez.write_bytes(b"%PDF-1.4")
    records = [{"Building": "John Stow House", "Brochure PDF": ""}]
    app_module._seed_brochure_from_source_pdf(
        records, breez, source_url, "rule:Breezblok", breez.name
    )
    assert records[0]["Brochure PDF"] == source_url
