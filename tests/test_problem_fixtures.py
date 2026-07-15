"""Regression coverage for the two 14 July problem emails."""
from pathlib import Path

from openpyxl import load_workbook

from extraction import pipeline
from extraction import geocode as geocode_module
from extraction.file_readers import read_file
from extraction.rules import try_rules
from extraction.schema import COLUMNS, normalize_record
from spreadsheet import write_xlsx


ROOT = Path(__file__).parent.parent
METSPACE = ROOT / "Fw_ MetSpace - Office Of The Week!.eml"
WORKPLACE = ROOT / "Fw_ Workplace Plus - Availability 14th July (1).eml"


def _records(path):
    rule, records = try_rules(read_file(path))
    return rule, [normalize_record(record) for record in records]


def test_metspace_office_of_week_assets_and_columns():
    rule, records = _records(METSPACE)
    assert rule == "MetSpace"
    assert len(records) == 1
    record = records[0]
    assert record["Building"] == "44 Pentonville Road"
    assert record["Size (sq ft)"] == 4413
    assert record["Desks (max)"] == 60
    assert record["Marketing Price (Based on Min Term) PCM"] == 42291
    assert record["Brochure PDF"].startswith("https://us.list-manage.com/")
    assert "457d7a07" in record["High Res Images"]  # verified property photograph
    assert "81074122" in record["Floor Plan"]  # verified floorplan diagram
    assert record["Floor Plan"] != record["High Res Images"]
    assert set(record).issuperset(COLUMNS)


def test_workplace_plus_rows_addresses_brochures_and_photos():
    rule, records = _records(WORKPLACE)
    assert rule == "Workplace Plus"
    assert [(r["Building"], r["Floor/Unit"], r["Desks (max)"], r["Marketing Price (Based on Min Term) PCM"]) for r in records] == [
        ("77 Gracechurch Street, EC3V 0AS", "6th Floor", 54, 55100),
        ("150 Waterloo Road, SE1 8SB", "2nd Floor", 24, 16000),
        ("150 Waterloo Road, SE1 8SB", "4th Floor", 18, 14600),
        ("8 Durweston Street, W1H 1EW", "Ground & First Floor", 24, 14600),
    ]
    assert [r["Property Postcode"] for r in records] == ["EC3V 0AS", "SE1 8SB", "SE1 8SB", "W1H 1EW"]
    assert all(r["Brochure PDF"].startswith("https://eot.workplaceplus.co.uk/") for r in records)
    assert all("gallery.eocampaign1.com" in r["High Res Images"] for r in records)
    assert all(not r["Floor Plan"] for r in records)
    assert records[1]["Brochure PDF"] == records[2]["Brochure PDF"]
    assert records[1]["High Res Images"] == records[2]["High Res Images"]
    forbidden = ("logo", "tentacles/icons", "linkedin", "website.png")
    assert all(not any(token in r["High Res Images"].lower() for token in forbidden) for r in records)


def test_exact_fixtures_run_through_orchestrator_and_qa(monkeypatch, tmp_path):
    postcode_by_query = {
        "44 Pentonville Road, London, UK": "N1 9HJ",
        "77 Gracechurch Street, EC3V 0AS, London, UK": "EC3V 0AS",
        "150 Waterloo Road, SE1 8SB, London, UK": "SE1 8SB",
        "8 Durweston Street, W1H 1EW, London, UK": "W1H 1EW",
    }

    def fake_geocode(query, confident=True):
        postcode = postcode_by_query.get(query)
        return (51.5, -0.1, postcode, None) if postcode else (None, None, None, "not found")

    monkeypatch.setattr(pipeline, "geocode", fake_geocode)
    source_urls = {
        METSPACE.name: "https://files.example.test/regression/metspace.eml",
        WORKPLACE.name: "https://files.example.test/regression/workplace-plus.eml",
    }
    results = pipeline.process_files([METSPACE, WORKPLACE], source_urls=source_urls, source_url_expected=True)
    assert [result["status"] for result in results] == ["ok", "ok"]
    assert [result["record_count"] for result in results] == [1, 4]
    assert results[0]["records"][0]["Property Postcode"] == "N1 9HJ (Not in source text)"

    # One-row and one-to-many extraction both retain exactly the originating
    # file reference; no listing asset can be substituted for provenance.
    for result, source in zip(results, (METSPACE, WORKPLACE)):
        expected_url = source_urls[source.name]
        assert all(prop.source_file_name == source.name for prop in result["properties"])
        assert all(prop.source_file_url == expected_url for prop in result["properties"])
        assert all(record["Link to File"] == expected_url for record in result["records"])
        assert all(record["_source_file_name"] == source.name for record in result["records"])
        assert all(
            record["Link to File"]
            not in {record["Brochure PDF"], record["Floor Plan"], record["High Res Images"]}
            for record in result["records"]
        )

    for result in results:
        path = tmp_path / f"{result['provider_name']}.xlsx"
        write_xlsx(path, result["records"])
        workbook = load_workbook(path)
        listings = workbook[workbook.sheetnames[0]]
        assert [cell.value for cell in listings[1]] == COLUMNS
        assert listings.max_column == len(COLUMNS)
        assert listings.max_row == result["record_count"] + 1
        # User-facing exports are Listings-only (QA Review is opt-in).
        assert "QA Review" not in workbook.sheetnames
        link_column = COLUMNS.index("Link to File") + 1
        source_name = METSPACE.name if result["provider_name"] == "MetSpace" else WORKPLACE.name
        for row in range(2, listings.max_row + 1):
            assert listings.cell(row, link_column).value == source_name
            assert listings.cell(row, link_column).hyperlink is not None

    by_provider = {result["provider_name"]: result for result in results}
    write_xlsx(tmp_path / "MetSpace-qa.xlsx", by_provider["MetSpace"]["records"], include_qa_sheet=True)
    metspace_qa = load_workbook(tmp_path / "MetSpace-qa.xlsx")["QA Review"]
    metspace_issues = [cell.value for cell in metspace_qa["D"]]
    assert any("derived during enrichment" in str(value) for value in metspace_issues)
    write_xlsx(tmp_path / "Workplace Plus-qa.xlsx", by_provider["Workplace Plus"]["records"], include_qa_sheet=True)
    workplace_qa = load_workbook(tmp_path / "Workplace Plus-qa.xlsx")["QA Review"]
    workplace_fields = [cell.value for cell in workplace_qa["C"]]
    assert workplace_fields.count("Size (sq ft)") == 4


def test_geocoder_rejects_neighbouring_lettered_unit():
    assert geocode_module._requested_house_number("44 Pentonville Road, London, UK") == "44"
    assert geocode_module._candidate_matches_house_number({"address": {"house_number": "44"}}, "44")
    assert not geocode_module._candidate_matches_house_number({"address": {"house_number": "44A"}}, "44")
    assert geocode_module._candidate_matches_house_number({"address": {"house_number": "42-44"}}, "44")
