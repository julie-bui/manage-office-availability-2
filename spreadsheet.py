"""Writes a single source's extracted records to a formatted .xlsx."""
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from extraction.models import Severity
from extraction.schema import COLUMNS

LAT_COL_IDX = COLUMNS.index("Lat") + 1

HEADER_FILL = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
CURRENCY_COLS = {"Marketing Price (Based on Min Term) PCM", "Marketing Price (Based on Min Term) PSF"}
NUMBER_COLS = {"Size (sq ft)", "Desks (max)"}
COORDINATE_COLS = {"Lat", "Lng"}
# Columns whose value is a URL, shown as a short clickable label instead
# of the raw link text (e.g. a long UUID-based URL) — same pattern for
# all four: cell.value becomes the label, cell.hyperlink keeps the real
# URL. Left blank (not the label) when there's genuinely no URL for that
# row.
LINK_LABELS = {
    "Brochure PDF": "Here",
    "Floor Plan": "Floor Plan",
    "High Res Images": "High Res Images",
}
# Every data cell wraps long text within its column (capped width below)
# rather than spilling sideways; row height is grown to fit after widths
# are finalized. Hyperlink columns keep short display labels so they stay
# compact.
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
WRAP_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
LINE_HEIGHT = 15  # approx. points needed per wrapped line at 11pt Calibri
MAX_COL_WIDTH = 45


def _estimate_wrapped_lines(text, width):
    """Estimate how many display lines wrapped text needs at the given column width."""
    if not text:
        return 1
    col_width = max(width or 10, 1)
    lines = 0
    for paragraph in str(text).splitlines() or [""]:
        # Empty paragraph (from a trailing newline / blank line) still takes a row.
        lines += max(1, math.ceil(len(paragraph) / col_width)) if paragraph else 1
    return max(lines, 1)


def _apply_row_heights(ws, start_row, end_row, col_count):
    """Set explicit row heights so wrap_text is visible without Excel AutoFit."""
    for row_idx in range(start_row, end_row + 1):
        max_lines = 1
        for col_idx in range(1, col_count + 1):
            letter = get_column_letter(col_idx)
            cell = ws.cell(row=row_idx, column=col_idx)
            text = str(cell.value) if cell.value is not None else ""
            if not text:
                continue
            width = ws.column_dimensions[letter].width or 10
            max_lines = max(max_lines, _estimate_wrapped_lines(text, width))
        if max_lines > 1:
            ws.row_dimensions[row_idx].height = max_lines * LINE_HEIGHT


def write_xlsx(path, records, sheet_title="Listings", include_qa_sheet=False):
    """Write the Listings sheet. QA Review is opt-in only (include_qa_sheet=True);
    user downloads omit it so the export is a single clean listings workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Listings"  # Excel sheet name length limit

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
    ws.freeze_panes = "A2"

    for row_idx, record in enumerate(records, start=2):
        row = [record.get(c, "") for c in COLUMNS]
        ws.append(row)
        # extraction.pipeline stashes this whenever Lat/Lng/Property
        # Postcode were derived via the web-search fallback (not read from
        # the source) — surfaced here as a cell comment so the specific
        # sources a "(Not in source text)" value was based on are visible
        # directly in the spreadsheet, not just the console log.
        sources = record.get("_geocode_sources")
        if sources:
            comment_text = "Address found via web search, based on:\n" + "\n".join(sources)
            ws.cell(row=row_idx, column=LAT_COL_IDX).comment = Comment(comment_text, "manage-office-availability")

    last_row = ws.max_row
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for row_idx in range(2, last_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            cell.alignment = WRAP_ALIGNMENT
            if col_name in CURRENCY_COLS and isinstance(val, (int, float)) and val != "":
                cell.number_format = "£#,##0.00" if col_name.endswith("PSF") else "£#,##0"
            elif col_name in NUMBER_COLS and isinstance(val, (int, float)) and val != "":
                cell.number_format = "#,##0"
            elif col_name in COORDINATE_COLS and isinstance(val, (int, float)) and val != "":
                cell.number_format = "0.000000"
            elif col_name in LINK_LABELS and isinstance(val, str) and val.startswith("http"):
                # Show a short label instead of the raw URL — the actual
                # link still goes to the real address via cell.hyperlink,
                # only the displayed text changes.
                actual_url = val
                cell.value = LINK_LABELS[col_name]
                cell.hyperlink = actual_url
                cell.font = Font(color="FF0563C1", underline="single")
                val = cell.value
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), MAX_COL_WIDTH)

    # wrap_text alone doesn't make Excel grow the row to fit — that's a
    # rendering computation Excel only does when a human triggers "AutoFit
    # Row Height", not on file load. So estimate wrapped line count from the
    # now-final column widths and set row height explicitly for every column.
    _apply_row_heights(ws, 2, last_row, len(COLUMNS))

    if include_qa_sheet:
        _write_qa_sheet(wb, records)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


QA_COLUMNS = ["Source File", "Property/Building", "Field", "Issue", "Severity", "Extracted Value", "Suggested Review Action"]


def _write_qa_sheet(workbook, records):
    """Add a stable, named-column review sheet without changing Listings."""
    qa = workbook.create_sheet("QA Review")
    qa.append(QA_COLUMNS)
    for cell in qa[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
    qa.freeze_panes = "A2"

    for record in records:
        issues = list(record.get("_validation_issues") or [])
        if not issues and record.get("_review_required"):
            issues = [{"field": "Record", "message": "Record was flagged for review.", "severity": "warning", "value": "", "action": "Review source values."}]
        for issue in issues:
            if isinstance(issue, dict):
                get = issue.get
            else:
                get = lambda key, default="": getattr(issue, key, default)
            severity = get("severity", "warning")
            if isinstance(severity, Severity):
                severity = severity.value
            qa.append(
                [
                    record.get("_source_file", ""),
                    record.get("Building", ""),
                    get("field"),
                    get("message"),
                    str(severity).upper(),
                    get("value"),
                    get("action"),
                ]
            )
        qa_diagnostic_statuses = {
            "LINK_IDENTITY_MATCH", "LINK_IDENTITY_PROBABLE_MATCH",
            "LINK_IDENTITY_AMBIGUOUS", "LINK_IDENTITY_HARD_CONFLICT",
            "NO_IMAGES_DISCOVERED", "IMAGES_DISCOVERED_BUT_REJECTED", "GALLERY_CREATION_FAILED",
            "LINK_EXPIRED_OR_INACCESSIBLE", "LINK_TIMED_OUT", "IMAGE_CANDIDATES_CLASSIFIED", "GALLERY_CREATED", "DIRECT_IMAGE_ASSIGNED",
        }
        for diagnostic in record.get("_link_diagnostics") or []:
            if isinstance(diagnostic, dict):
                get = diagnostic.get
            else:
                get = lambda key, default="": getattr(diagnostic, key, default)
            status = get("status", "LINK_DIAGNOSTIC")
            if status not in qa_diagnostic_statuses:
                continue
            detail = get("detail", "")
            identity = get("identity_result", "")
            qa.append(
                [
                    record.get("_source_file", ""),
                    record.get("Building", ""),
                    "Linked Media",
                    f"{status}: {detail}".rstrip(": "),
                    "WARNING" if status in {"LINK_IDENTITY_AMBIGUOUS", "LINK_IDENTITY_HARD_CONFLICT", "IMAGES_DISCOVERED_BUT_REJECTED", "GALLERY_CREATION_FAILED", "LINK_EXPIRED_OR_INACCESSIBLE", "LINK_TIMED_OUT"} else "INFO",
                    get("final_url", "") or get("original_url", ""),
                    f"Identity: {identity}" if identity else "No action required unless the linked media is missing or incorrect.",
                ]
            )

    if qa.max_row == 1:
        qa.append(["", "", "", "No validation issues detected", "INFO", "", "No action required"])
    for idx, name in enumerate(QA_COLUMNS, start=1):
        qa.column_dimensions[get_column_letter(idx)].width = min(max(len(name) + 2, 16), MAX_COL_WIDTH)
    for row in qa.iter_rows():
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT
    # Same as Listings: wrap_text needs an explicit height estimate so long
    # Issue / action text is visible without Excel AutoFit.
    _apply_row_heights(qa, 2, qa.max_row, len(QA_COLUMNS))
